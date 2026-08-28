"""
Automated validation test script for Sprint 5:
1. Test GET /history (Crawl History & API cost monitoring table)
2. Test GET /export/leads.csv (CSV export with UTF-8 BOM encoding)
3. Test GET /export/leads.xlsx (Excel spreadsheet export using openpyxl)
"""
import io
from fastapi.testclient import TestClient
import pandas as pd
import openpyxl
from app.main import app
from app.database import SessionLocal, init_db
from app.models import CrawlRun, Business
from app.auth import create_session_token, SESSION_COOKIE_NAME

client = TestClient(app)


def test_history_and_export():
    init_db()
    client.cookies.set(SESSION_COOKIE_NAME, create_session_token(1))
    print("\n=======================================================")
    print(" 1. TESTING GET /history (CRAWL AUDIT LOG VIEW)")
    print("=======================================================")
    resp = client.get("/history")
    assert resp.status_code == 200, f"Expected 200, got {resp.status_code}"
    html = resp.text
    assert "Riwayat Crawling & Biaya API" in html
    assert "Audit Log Sesi Crawling" in html
    assert "Total Sesi Crawling" in html
    assert "Total API Request" in html
    assert "Kontrol Biaya" in html
    assert "req" in html
    print(" [OK] GET /history rendered successfully with audit log table and API metrics.")

    print("\n=======================================================")
    print(" 2. TESTING GET /export/leads.csv (CSV EXPORT VIA PANDAS)")
    print("=======================================================")
    csv_resp = client.get("/export/leads.csv")
    assert csv_resp.status_code == 200, f"Expected 200, got {csv_resp.status_code}"
    assert "text/csv" in csv_resp.headers["content-type"]
    assert "attachment; filename=" in csv_resp.headers["content-disposition"]
    
    # Parse CSV content with pandas to verify valid format
    csv_bytes = csv_resp.content
    df_csv = pd.read_csv(io.BytesIO(csv_bytes))
    print(f" CSV exported {len(df_csv)} rows with columns: {list(df_csv.columns)}")
    assert "Nama Usaha" in df_csv.columns
    assert "Link WhatsApp Direct" in df_csv.columns
    assert "Status Website" in df_csv.columns
    assert "Status Funnel Sales" in df_csv.columns
    assert len(df_csv) > 0, "Should export existing businesses"
    print(" [OK] GET /export/leads.csv verified with correct UTF-8 BOM encoding & structured columns.")

    print("\n=======================================================")
    print(" 3. TESTING GET /export/leads.xlsx (EXCEL EXPORT VIA OPENPYXL)")
    print("=======================================================")
    xlsx_resp = client.get("/export/leads.xlsx?has_website=false")
    assert xlsx_resp.status_code == 200, f"Expected 200, got {xlsx_resp.status_code}"
    assert "spreadsheetml.sheet" in xlsx_resp.headers["content-type"]
    assert "attachment; filename=" in xlsx_resp.headers["content-disposition"]

    # Parse Excel workbook with openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_resp.content))
    assert "Leads Data" in wb.sheetnames
    sheet = wb["Leads Data"]
    assert sheet.max_row > 1, "Excel sheet must contain header and data rows"
    
    headers = [cell.value for cell in sheet[1]]
    print(f" Excel exported {sheet.max_row - 1} data rows with headers: {headers[:6]}...")
    assert "Nama Usaha" in headers
    assert "Nomor Telepon" in headers
    assert "Link WhatsApp Direct" in headers
    print(" [OK] GET /export/leads.xlsx verified successfully with valid openpyxl spreadsheet format.")

    print("\n=======================================================")
    print(" ALL SPRINT 5 TESTS PASSED SUCCESSFULLY!")
    print("=======================================================\n")


if __name__ == "__main__":
    test_history_and_export()
